"""
Generare Excel Activitate Clienti Zitamine
==========================================
Structura Excel:
- Client ID (Order Name al primei comenzi)
- Email
- Cohorta (SUB1, SUB3, SUB6, OTP)
- Tip Abonament: luna-anSUBx / luna-anOTP (ex: 1-2025SUB1 = client pornit in ian 2025 cu SUB1)
- Data Start (data primei comenzi din cohorta curenta)
- Coloane lunare (Ian 2024 -> luna curenta): tipul abonament activ sau 0
- Conversion Event (daca a schimbat tipul de abonament)
- Status Actual: ACTIV / INACTIV / DROPOUT

Reguli Status:
  ACTIV:   OTP/SUB1 < 3 luni | SUB3 < 5 luni | SUB6 < 7 luni (fara comanda)
  INACTIV: OTP/SUB1 3 luni   | SUB3 5 luni   | SUB6 7 luni
  DROPOUT: OTP/SUB1 >= 4 luni | SUB3 >= 6 luni | SUB6 >= 9 luni
"""

import csv
import re
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# === CONFIG ===
INPUT_FILE = r"c:\Users\Zitamine\Victor Dane\Antigravity\Proiecte-Varianta-Finala\Raport_Vanzari_2024_2026\master_orders.csv"
OUTPUT_FILE = r"c:\Users\Zitamine\Victor Dane\Antigravity\Cohort tracking + Raport Dropout.xlsx"

# Start month for columns (inclusive)
START_YEAR = 2024
START_MONTH = 1
# End month = current month
NOW = datetime.now()
END_YEAR = NOW.year
END_MONTH = NOW.month

def generate_month_columns():
    """Generate list of month keys from START to END."""
    months = []
    y, m = START_YEAR, START_MONTH
    while (y, m) <= (END_YEAR, END_MONTH):
        months.append(f"{y}-{m:02d}")
        m += 1
        if m > 12:
            m = 1
            y += 1
    return months

def month_label(month_key):
    """Convert '2024-01' to 'Ian 2024' etc."""
    labels = {1: 'Ian', 2: 'Feb', 3: 'Mar', 4: 'Apr', 5: 'Mai', 6: 'Iun',
              7: 'Iul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12: 'Dec'}
    y, m = month_key.split('-')
    return f"{labels[int(m)]} {y}"

def classify_order_type(tags):
    """Classify order type based on tags (matching PowerShell logic)."""
    if not tags:
        return "OTP"
    tags_lower = tags.lower()
    if "saseluni" in tags_lower:
        return "SUB6"
    elif "treiluni" in tags_lower or "treluni" in tags_lower:
        return "SUB3"
    elif "appstle_subscription" in tags_lower:
        return "SUB1"
    elif "subscription" in tags_lower:
        return "SUB1"
    return "OTP"

def is_subscription_tag(tags):
    """Check if order has subscription or OTP tags (i.e., is a tracked order)."""
    if not tags:
        return False
    tags_lower = tags.lower()
    return any(t in tags_lower for t in [
        'appstle_subscription_first_order',
        'appstle_subscription_recurring_order',
        'otp_first_order', 'otp first order', 'otp_first_oder',
        'otp_recurring_order', 'otp_recurring order', 'ramburs otp_recurring'
    ])

def parse_date(date_str):
    """Parse date string from CSV."""
    if not date_str:
        return None
    try:
        # Remove timezone info
        clean = date_str.split('+')[0].strip()
        if 'T' in clean:
            return datetime.strptime(clean, '%Y-%m-%dT%H:%M:%S')
        else:
            return datetime.strptime(clean, '%Y-%m-%d')
    except:
        return None

def tip_abonament_label(month_key, order_type):
    """Generate label like '1-2025SUB1' or '3-2026OTP'."""
    y, m = month_key.split('-')
    return f"{int(m)}-{y}{order_type}"

def main():
    print("Loading master_orders.csv...")
    
    # Step 1: Read all orders (deduplicate by order name, aggregate per customer)
    orders_by_name = {}  # order_name -> {email, date, tags, order_type, ...}
    
    with open(INPUT_FILE, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            order_name = row.get('Name', '').strip()
            if not order_name:
                continue
            
            # Skip if already processed this order
            if order_name in orders_by_name:
                continue
            
            email = row.get('Email', '').strip().lower()
            if not email:
                continue
            
            # Skip cancelled orders
            cancelled = row.get('Cancelled at', '').strip()
            if cancelled:
                continue
            
            # Skip pending
            fin_status = (row.get('Financial Status', '') or '').strip().lower()
            if fin_status == 'pending':
                continue
            
            date_obj = parse_date(row.get('Created at', ''))
            if not date_obj:
                continue
            
            tags = row.get('Tags', '').strip()
            order_type = classify_order_type(tags)
            month_key = f"{date_obj.year}-{date_obj.month:02d}"
            
            orders_by_name[order_name] = {
                'email': email,
                'date': date_obj,
                'tags': tags,
                'order_type': order_type,
                'month_key': month_key,
                'order_name': order_name,
            }
    
    print(f"Total unique valid orders: {len(orders_by_name)}")
    
    # Step 2: Group orders by customer email
    customer_orders = defaultdict(list)
    for order_name, order in orders_by_name.items():
        customer_orders[order['email']].append(order)
    
    # Sort each customer's orders by date
    for email in customer_orders:
        customer_orders[email].sort(key=lambda x: x['date'])
    
    print(f"Total unique customers: {len(customer_orders)}")
    
    # Step 3: Build the activity matrix
    all_months = generate_month_columns()
    print(f"Month columns: {len(all_months)} (from {all_months[0]} to {all_months[-1]})")
    
    rows_data = []
    
    for email, orders in sorted(customer_orders.items(), key=lambda x: x[1][0]['date']):
        first_order = orders[0]
        first_order_date = first_order['date']
        first_order_type = first_order['order_type']
        first_month_key = first_order['month_key']
        
        # Client ID = first order name
        client_id = first_order['order_name']
        
        # Current cohort = type of the most recent order
        current_cohort = orders[-1]['order_type']
        
        # Tip Abonament = based on first order month + first order type
        tip_abonament = tip_abonament_label(first_month_key, first_order_type)
        
        # Data Start
        data_start = first_order_date.strftime('%d/%m/%Y')
        
        # Build monthly activity
        # Map each month to the orders placed in that month
        month_to_orders = defaultdict(list)
        for order in orders:
            month_to_orders[order['month_key']].append(order)
        
        # Track the "current type" as it evolves through months
        monthly_activity = {}
        current_type_tracker = first_order_type
        
        # Conversion events
        conversions = []
        
        # Track the starting month of each subscription segment
        current_segment_start_month = int(first_month_key.split('-')[1])
        
        for month in all_months:
            if month in month_to_orders:
                # Customer has order(s) this month
                month_orders = month_to_orders[month]
                # Use the last order's type for this month (in case of multiple)
                month_type = month_orders[-1]['order_type']
                
                # Check for conversion
                if current_type_tracker != month_type:
                    # Conversion happened - update segment start month
                    y, m = month.split('-')
                    month_names = {1:'Ian',2:'Feb',3:'Mar',4:'Apr',5:'Mai',6:'Iun',
                                   7:'Iul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}
                    conversion_label = f"{month_names[int(m)]}-{y[-2:]}: {current_type_tracker}→{month_type}"
                    conversions.append(conversion_label)
                    current_type_tracker = month_type
                    current_segment_start_month = int(m)
                
                # Label: segment start month + current type (e.g., 1SUB1, 2OTP)
                monthly_activity[month] = f"{current_segment_start_month}{month_type}"
            else:
                # No order this month
                monthly_activity[month] = 0
        
        conversion_text = "; ".join(conversions) if conversions else ""
        
        # === STATUS ACTUAL ===
        # Calculate months since last order
        last_order_date = orders[-1]['date']
        last_order_type = orders[-1]['order_type']
        
        # Month diff from last order to current month
        months_since_last = (NOW.year - last_order_date.year) * 12 + (NOW.month - last_order_date.month)
        
        # Inactiv thresholds (existing churn rules)
        inactiv_thresholds = {'OTP': 3, 'SUB1': 3, 'SUB3': 5, 'SUB6': 7}
        # Dropout thresholds (new rules)
        dropout_thresholds = {'OTP': 4, 'SUB1': 4, 'SUB3': 6, 'SUB6': 9}
        
        inactiv_limit = inactiv_thresholds.get(last_order_type, 3)
        dropout_limit = dropout_thresholds.get(last_order_type, 4)
        
        if months_since_last >= dropout_limit:
            status_actual = 'DROPOUT'
        elif months_since_last >= inactiv_limit:
            status_actual = 'INACTIV'
        else:
            status_actual = 'ACTIV'
        
        rows_data.append({
            'client_id': client_id,
            'email': email,
            'cohort': current_cohort,
            'tip_abonament': tip_abonament,
            'data_start': data_start,
            'monthly': monthly_activity,
            'conversion': conversion_text,
            'status': status_actual,
        })
    
    print(f"Total customer rows: {len(rows_data)}")
    
    # Step 4: Write to Excel
    print("Writing Excel...")
    wb = Workbook()
    ws = wb.active
    ws.title = "Activitate Clienti"
    
    # === STYLES ===
    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    data_font = Font(name='Calibri', size=10)
    data_alignment = Alignment(horizontal='center', vertical='center')
    
    # Colors for subscription types
    type_fills = {
        'SUB1': PatternFill(start_color='3498DB', end_color='3498DB', fill_type='solid'),  # Blue
        'SUB3': PatternFill(start_color='E67E22', end_color='E67E22', fill_type='solid'),  # Orange
        'SUB6': PatternFill(start_color='9B59B6', end_color='9B59B6', fill_type='solid'),  # Purple
        'OTP':  PatternFill(start_color='F39C12', end_color='F39C12', fill_type='solid'),   # Yellow-Orange
    }
    type_fonts = {
        'SUB1': Font(name='Calibri', size=10, color='FFFFFF', bold=True),
        'SUB3': Font(name='Calibri', size=10, color='FFFFFF', bold=True),
        'SUB6': Font(name='Calibri', size=10, color='FFFFFF', bold=True),
        'OTP':  Font(name='Calibri', size=10, color='FFFFFF', bold=True),
    }
    
    zero_fill = PatternFill(start_color='F5F5F5', end_color='F5F5F5', fill_type='solid')
    zero_font = Font(name='Calibri', size=10, color='BDBDBD')
    
    conversion_fill = PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid')
    conversion_font = Font(name='Calibri', size=10, color='FFFFFF', bold=True)
    
    # Status styles
    status_fills = {
        'ACTIV':   PatternFill(start_color='27AE60', end_color='27AE60', fill_type='solid'),   # Green
        'INACTIV': PatternFill(start_color='F39C12', end_color='F39C12', fill_type='solid'),   # Orange
        'DROPOUT': PatternFill(start_color='E74C3C', end_color='E74C3C', fill_type='solid'),   # Red
    }
    status_font = Font(name='Calibri', size=10, color='FFFFFF', bold=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D5D8DC'),
        right=Side(style='thin', color='D5D8DC'),
        top=Side(style='thin', color='D5D8DC'),
        bottom=Side(style='thin', color='D5D8DC')
    )
    
    # === HEADERS ===
    headers = ['Client ID', 'Email', 'Cohorta', 'Tip Abonament', 'Data Start']
    month_headers = [month_label(m) for m in all_months]
    headers += month_headers
    headers.append('Conversion Event')
    headers.append('Status Actual')
    
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # === DATA ROWS ===
    for row_idx, data in enumerate(rows_data, 2):
        # Client ID
        cell = ws.cell(row=row_idx, column=1, value=data['client_id'])
        cell.font = data_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border
        
        # Email
        cell = ws.cell(row=row_idx, column=2, value=data['email'])
        cell.font = data_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = thin_border
        
        # Cohort
        cohort = data['cohort']
        cell = ws.cell(row=row_idx, column=3, value=cohort)
        cell.font = type_fonts.get(cohort, data_font)
        cell.fill = type_fills.get(cohort, PatternFill())
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Tip Abonament
        cell = ws.cell(row=row_idx, column=4, value=data['tip_abonament'])
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Data Start
        cell = ws.cell(row=row_idx, column=5, value=data['data_start'])
        cell.font = data_font
        cell.alignment = data_alignment
        cell.border = thin_border
        
        # Monthly columns
        for m_idx, month_key in enumerate(all_months):
            col = 6 + m_idx
            value = data['monthly'].get(month_key, 0)
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.alignment = data_alignment
            cell.border = thin_border
            
            if value == 0:
                cell.font = zero_font
                cell.fill = zero_fill
            else:
                # Determine the type from the value (e.g., "1SUB1" -> SUB1)
                value_str = str(value)
                matched_type = None
                for t in ['SUB6', 'SUB3', 'SUB1', 'OTP']:
                    if t in value_str:
                        matched_type = t
                        break
                if matched_type:
                    cell.font = type_fonts[matched_type]
                    cell.fill = type_fills[matched_type]
                else:
                    cell.font = data_font
        
        # Conversion Event
        conv_col = 6 + len(all_months)
        cell = ws.cell(row=row_idx, column=conv_col, value=data['conversion'])
        cell.border = thin_border
        if data['conversion']:
            cell.font = conversion_font
            cell.fill = conversion_fill
        else:
            cell.font = data_font
        cell.alignment = Alignment(horizontal='left', vertical='center')
        
        # Status Actual
        status_col = conv_col + 1
        status_val = data['status']
        cell = ws.cell(row=row_idx, column=status_col, value=status_val)
        cell.font = status_font
        cell.fill = status_fills.get(status_val, PatternFill())
        cell.alignment = data_alignment
        cell.border = thin_border
    
    # === COLUMN WIDTHS ===
    ws.column_dimensions['A'].width = 12  # Client ID
    ws.column_dimensions['B'].width = 30  # Email
    ws.column_dimensions['C'].width = 10  # Cohort
    ws.column_dimensions['D'].width = 16  # Tip Abonament
    ws.column_dimensions['E'].width = 12  # Data Start
    
    # Month columns
    for i in range(len(all_months)):
        col_letter = get_column_letter(6 + i)
        ws.column_dimensions[col_letter].width = 10
    
    # Conversion column
    conv_letter = get_column_letter(6 + len(all_months))
    ws.column_dimensions[conv_letter].width = 25
    
    # Status column
    status_letter = get_column_letter(6 + len(all_months) + 1)
    ws.column_dimensions[status_letter].width = 14
    
    # Freeze panes at row 2, column 6 (so headers + first 5 cols are frozen)
    ws.freeze_panes = 'F2'
    
    # Auto-filter
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(rows_data)+1}"
    
    # === ADD LEGENDA SHEET ===
    ws_legenda = wb.create_sheet(title="Legenda")
    
    # Legend data
    legend_content = [
        ["LEGENDĂ - URMĂRIRE ACTIVITATE CLIENȚI", ""],
        ["", ""],
        ["1. STATUS ACTUAL (Ultima Coloană)", "Definiția și pragurile de inactivitate/dropout pe bază de cohorte."],
        ["Activ", "Client activ, care plasează comenzi în mod regulat. Nu a depășit pragul de inactivitate."],
        ["Inactiv", "Client aflat la limită, fără comandă curentă. Riscă să devină dropout: \nOTP / SUB1 = Exact 3 luni fără comandă \nSUB3 = Exact 5 luni fără comandă \nSUB6 = Exact 7 luni fără comandă"],
        ["Dropout", "Client pierdut. Pragul la care un client este declarat dropout: \nOTP / SUB1 = 4+ luni fără comandă \nSUB3 = 6+ luni fără comandă \nSUB6 = 9+ luni fără comandă"],
        ["", ""],
        ["2. MATRICEA LUNARĂ DE ACTIVITATE", "Codificarea celulelor din fiecare lună."],
        ["0", "Nu a plasat nicio comandă în luna respectivă."],
        ["1SUB1, 2SUB1...", "Exemplu: Dacă scrie '1SUB1', înseamnă că abonamentul SUB1 actual are data de start în LUNA 1 (Ianuarie). Cifra indică LUNA DE START, iar textul indică TIPUL de abonament (SUB 1 lună)."],
        ["1OTP, 3OTP...", "Exemplu: '3OTP' înseamnă că ultima comandă curentă a început un ciclu în luna 3 (Martie), iar tipul este plată One-Time Purchase (fără abonament)."],
        ["1SUB3, 5SUB3...", "Abonament recurent la 3 luni (SUB3), ciclul curent a început în luna indicată de cifra prefixului."],
        ["Conversie", "Când prefixul se schimbă, clientul a tranzitat. \nEx: În Ian. este '1SUB1', iar în Martie este '3OTP'. Înseamnă că în Ian a cumpărat SUB1, și în Martie a devenit OTP. Acest eveniment se notează pe coloana de conversie."],
    ]
    
    for row_idx, row_data in enumerate(legend_content, 1):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws_legenda.cell(row=row_idx, column=col_idx, value=value)
            # Styling definitions
            if "LEGENDĂ" in str(value):
                cell.font = Font(bold=True, size=14)
            elif "1." in str(value) or "2." in str(value):
                cell.font = Font(bold=True, size=12)
                cell.fill = PatternFill(start_color='E0E0E0', end_color='E0E0E0', fill_type='solid')
            if col_idx == 1:
                cell.font = Font(bold=True)
            cell.alignment = Alignment(vertical='top', wrap_text=True)
    
    ws_legenda.column_dimensions['A'].width = 30
    ws_legenda.column_dimensions['B'].width = 80
    
    # Save
    wb.save(OUTPUT_FILE)
    print(f"\nExcel saved to: {OUTPUT_FILE}")
    print(f"Total rows: {len(rows_data)}")
    print(f"Total columns: {len(headers)}")

if __name__ == "__main__":
    main()
