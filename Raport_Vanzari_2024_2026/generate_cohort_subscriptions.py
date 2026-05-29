"""Raport Cohorte Abonamente Zitamine — Varianta "Subscription First Order"

Diferență față de raportul principal:
- OTP: la fel (clienți 100% noi, prima comandă ever)
- SUB1/SUB3/SUB6: "client nou" = oricine cu tag appstle_subscription_first_order
  (inclusiv foști clienți OTP care trec pe abonament)
"""
import csv, os
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_CSV = os.path.join(SCRIPT_DIR, "master_orders.csv")
OUTPUT_EXCEL = os.path.join(SCRIPT_DIR, "..", "..", "Rapoarte", "Raport_Cohorte_Abonamente.xlsx")

COHORT_START, COHORT_END = "2025-05", "2026-03"
MILESTONES = [1, 2, 3, 6, 9, 12]
MILESTONE_RANGES = {1: (0,1), 2: (2,2), 3: (3,3), 6: (4,6), 9: (7,9), 12: (10,12)}
MILESTONE_LABELS = {1:'M0-1', 2:'M2', 3:'M3', 6:'M4-6', 9:'M7-9', 12:'M10-12'}
REPORT_DATE = datetime(2026, 4, 30)
ORDER_TYPES = ['OTP', 'SUB1', 'SUB3', 'SUB6']
TYPE_COLORS = {'OTP':'#F39C12','SUB1':'#3498DB','SUB3':'#E67E22','SUB6':'#9B59B6'}
TYPE_DESC = {
    'OTP': 'One-Time Purchase — Clienți 100% noi',
    'SUB1': 'Abonament Lunar — Noi abonați (incl. foști OTP)',
    'SUB3': 'Abonament 3 Luni — Noi abonați (incl. foști OTP)',
    'SUB6': 'Protocol 6 Luni — Noi abonați (incl. foști OTP)',
}
ML = {1:'Ian',2:'Feb',3:'Mar',4:'Apr',5:'Mai',6:'Iun',7:'Iul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}

def target_months():
    r, y, m = [], 2025, 5
    while f"{y}-{m:02d}" <= COHORT_END:
        r.append(f"{y}-{m:02d}"); m += 1
        if m > 12: m, y = 1, y+1
    return r

def mlabel(mk):
    y, m = mk.split('-'); return f"{ML[int(m)]} {y}"

def parse_date(s):
    if not s: return None
    try:
        c = s.split('+')[0].strip()
        return datetime.strptime(c, '%Y-%m-%dT%H:%M:%S') if 'T' in c else datetime.strptime(c, '%Y-%m-%d')
    except: return None

def classify_type(tags):
    if not tags: return "OTP"
    t = tags.lower()
    if "saseluni" in t: return "SUB6"
    if "treiluni" in t or "treluni" in t: return "SUB3"
    if "appstle_subscription" in t or "subscription" in t: return "SUB1"
    return "OTP"

def mdiff(d1, d2):
    return (d1.year - d2.year)*12 + (d1.month - d2.month)

def load_orders():
    print("Loading master_orders.csv...")
    orders = {}
    with open(MASTER_CSV, 'r', encoding='utf-8-sig') as f:
        for row in csv.DictReader(f):
            name = row.get('Name','').strip()
            if not name or name in orders: continue
            email = row.get('Email','').strip().lower()
            if not email: continue
            if row.get('Cancelled at','').strip(): continue
            fs = (row.get('Financial Status','') or '').strip().lower()
            if fs in ('voided','pending'): continue
            d = parse_date(row.get('Created at',''))
            if not d: continue
            try: total = float(row.get('Total', 0) or 0)
            except: total = 0
            tags = row.get('Tags', '') or ''
            orders[name] = {
                'email': email, 'date': d, 'total': total,
                'month': f"{d.year}-{d.month:02d}",
                'type': classify_type(tags),
                'tags': tags.lower(),
                'is_sub_first': 'appstle_subscription_first_order' in tags.lower(),
            }
    print(f"  {len(orders)} valid orders")
    return orders


def build_cohorts(orders):
    """
    OTP cohort: client nou = prima comandă ever (la fel ca raportul principal).
    SUB cohorts: client nou = oricine cu appstle_subscription_first_order
                 (chiar dacă a fost client OTP înainte).
    """
    tm = target_months()

    # Group orders by email
    cust = defaultdict(list)
    for o in orders.values():
        cust[o['email']].append(o)
    for e in cust:
        cust[e].sort(key=lambda x: x['date'])

    # Init cohorts
    cohorts = {}
    for ot in ORDER_TYPES + ['ALL']:
        cohorts[ot] = {}
        for mk in tm:
            cohorts[ot][mk] = {
                'label': mlabel(mk), 'customers': 0, 'total_rev': 0,
                'retained': {ms: 0 for ms in MILESTONES},
                'pct': {ms: None for ms in MILESTONES}
            }

    # === OTP COHORT (same as main report: first order ever) ===
    for email, ords in cust.items():
        first = ords[0]
        cmk = first['month']
        if cmk not in cohorts['ALL']: continue
        if first['type'] != 'OTP': continue

        cohorts['OTP'][cmk]['customers'] += 1
        cohorts['OTP'][cmk]['total_rev'] += sum(o['total'] for o in ords)
        if len(ords) >= 2:
            m_to_2nd = mdiff(ords[1]['date'], first['date'])
            for ms in MILESTONES:
                lo, hi = MILESTONE_RANGES[ms]
                if lo <= m_to_2nd <= hi:
                    cohorts['OTP'][cmk]['retained'][ms] += 1

    # === SUB COHORTS (subscription_first_order = new subscriber) ===
    for email, ords in cust.items():
        # Find all sub_first orders for this customer
        sub_firsts = [o for o in ords if o['is_sub_first']]
        for sf in sub_firsts:
            cmk = sf['month']
            if cmk not in cohorts['ALL']: continue
            sub_type = sf['type']  # SUB1, SUB3, or SUB6

            # Add to specific sub cohort
            cohorts[sub_type][cmk]['customers'] += 1

            # Revenue: sum all orders AFTER (and including) this sub first order
            subsequent = [o for o in ords if o['date'] >= sf['date']]
            cohorts[sub_type][cmk]['total_rev'] += sum(o['total'] for o in subsequent)

            # Retention: find the NEXT order after this sub first
            later_orders = [o for o in ords if o['date'] > sf['date']]
            if later_orders:
                next_order = later_orders[0]
                m_to_next = mdiff(next_order['date'], sf['date'])
                for ms in MILESTONES:
                    lo, hi = MILESTONE_RANGES[ms]
                    if lo <= m_to_next <= hi:
                        cohorts[sub_type][cmk]['retained'][ms] += 1

    # === ALL COHORT (first order ever, any type — same as main report) ===
    for email, ords in cust.items():
        first = ords[0]
        cmk = first['month']
        if cmk not in cohorts['ALL']: continue
        cohorts['ALL'][cmk]['customers'] += 1
        cohorts['ALL'][cmk]['total_rev'] += sum(o['total'] for o in ords)
        if len(ords) >= 2:
            m_to_2nd = mdiff(ords[1]['date'], first['date'])
            for ms in MILESTONES:
                lo, hi = MILESTONE_RANGES[ms]
                if lo <= m_to_2nd <= hi:
                    cohorts['ALL'][cmk]['retained'][ms] += 1

    # Calculate percentages
    for ot in list(cohorts.keys()):
        for mk, c in cohorts[ot].items():
            avail = mdiff(REPORT_DATE, datetime(*map(int, mk.split('-')), 1))
            for ms in MILESTONES:
                if avail >= ms and c['customers'] > 0:
                    c['pct'][ms] = round(c['retained'][ms] / c['customers'] * 100, 1)
            c['ltv'] = round(c['total_rev'] / c['customers'], 2) if c['customers'] > 0 else 0

    for ot in ORDER_TYPES + ['ALL']:
        total = sum(cohorts[ot][mk]['customers'] for mk in tm)
        print(f"  {ot}: {total} customers across {len(tm)} months")
    return cohorts


def gen_excel(cohorts, traffic=None):
    tm = target_months()
    wb = Workbook()

    # Styles
    hf = Font(name='Inter', bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='1a1f36', end_color='1a1f36', fill_type='solid')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    df = Font(name='Inter', size=10)
    da = Alignment(horizontal='center', vertical='center')
    brd = Border(left=Side('thin','D5D8DC'), right=Side('thin','D5D8DC'),
                 top=Side('thin','D5D8DC'), bottom=Side('thin','D5D8DC'))
    na_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    na_font = Font(name='Inter', size=10, color='BDBDBD', italic=True)
    gf = [PatternFill(start_color=c, end_color=c, fill_type='solid')
          for c in ['E8F5E9','C8E6C9','A5D6A7','81C784','66BB6A','4CAF50']]

    def sh(ws, row, cols):
        for c in range(1, cols+1):
            cell = ws.cell(row=row, column=c)
            cell.font, cell.fill, cell.alignment, cell.border = hf, hfill, ha, brd

    def sd(cell):
        cell.font, cell.alignment, cell.border = df, da, brd

    for si, ot in enumerate(['ALL'] + ORDER_TYPES):
        if si == 0:
            ws = wb.active; ws.title = "ALL - Total"
        else:
            ws = wb.create_sheet(title=ot)

        # Determine column header based on type
        if ot == 'OTP':
            cust_header = '# Clienți Noi\n(prima comandă ever)'
        elif ot == 'ALL':
            cust_header = '# Clienți Noi\n(prima comandă ever)'
        else:
            cust_header = '# Noi Abonați\n(subscription_first_order)'

        headers = ['Cohortă', cust_header] + [f'% R2 {MILESTONE_LABELS[ms]}' for ms in MILESTONES] + ['LTV (RON)']
        ncols = len(headers)

        # Title
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        type_label = f"{ot} — {TYPE_DESC[ot]}" if ot in TYPE_DESC else 'TOTAL (toate tipurile — prima comandă ever)'
        title_cell = ws.cell(row=1, column=1, value=f"Cohort Retention & LTV — {type_label}")
        title_cell.font = Font(name='Inter', bold=True, size=13, color='1a1f36')
        title_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Sub-title explaining methodology
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=ncols)
        if ot in ('SUB1', 'SUB3', 'SUB6'):
            subtitle = "Client nou = oricine cu tag appstle_subscription_first_order (inclusiv fosti OTP convertiti la abonament)"
        elif ot == 'OTP':
            subtitle = "Client nou = prima comandă ever la Zitamine (nicio comandă anterioară)"
        else:
            subtitle = "Client nou = prima comandă ever la Zitamine (orice tip)"
        sub_cell = ws.cell(row=2, column=1, value=subtitle)
        sub_cell.font = Font(name='Inter', size=9, italic=True, color='697386')
        sub_cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        # Headers at row 3
        for ci, h in enumerate(headers, 1):
            ws.cell(row=3, column=ci, value=h)
        sh(ws, 3, ncols)

        for ri, mk in enumerate(tm, 4):
            c = cohorts[ot][mk]
            ws.cell(row=ri, column=1, value=c['label']); sd(ws.cell(row=ri, column=1))
            ws.cell(row=ri, column=1).font = Font(name='Inter', size=10, bold=True)
            ws.cell(row=ri, column=2, value=c['customers']); sd(ws.cell(row=ri, column=2))
            for mi, ms in enumerate(MILESTONES):
                col = 3 + mi
                pct = c['pct'][ms]
                cell = ws.cell(row=ri, column=col)
                if pct is not None:
                    cell.value = pct / 100; cell.number_format = '0.0%'; sd(cell)
                    if pct >= 30: cell.fill = gf[min(int(pct/10)-2, 5)]
                    elif pct >= 15: cell.fill = gf[1]
                    elif pct > 0: cell.fill = gf[0]
                else:
                    cell.value = 'N/A'; cell.font, cell.fill = na_font, na_fill
                    cell.alignment, cell.border = da, brd
            ltv_cell = ws.cell(row=ri, column=ncols, value=c['ltv'])
            sd(ltv_cell); ltv_cell.number_format = '#,##0.00 "RON"'

        for ci, w in enumerate([16, 18] + [13]*len(MILESTONES) + [16], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = 'C4'

    # Traffic sheet
    ws2 = wb.create_sheet("Revenue by Traffic Source")
    cats = ['organic','paid','direct','email','other']
    clab = {'organic':'Organic','paid':'Paid','direct':'Direct','email':'Email (Klaviyo)','other':'Other'}
    h2 = ['Luna', 'Total (RON)']
    for cat in cats: h2 += [f'{clab[cat]} (RON)', f'% {clab[cat]}']
    h2.append('Precizie Date')
    for ci, h in enumerate(h2, 1): ws2.cell(row=1, column=ci, value=h)
    sh(ws2, 1, len(h2))
    cfills = {'organic':'E8F5E9','paid':'FFF3E0','direct':'E3F2FD','email':'F3E5F5','other':'F5F5F5'}
    est_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')  # yellow
    precise_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')  # green

    if traffic:
        for ri, mk in enumerate(tm, 2):
            t = traffic.get(mk, {k:0 for k in cats+['total']})
            ws2.cell(row=ri, column=1, value=mlabel(mk)); sd(ws2.cell(row=ri, column=1))
            ws2.cell(row=ri, column=1).font = Font(name='Inter', size=10, bold=True)
            ws2.cell(row=ri, column=2, value=round(t['total'],2)); sd(ws2.cell(row=ri, column=2))
            ws2.cell(row=ri, column=2).number_format = '#,##0.00'
            for ci2, cat in enumerate(cats):
                fill = PatternFill(start_color=cfills[cat], end_color=cfills[cat], fill_type='solid')
                rev = t[cat]; pct = rev/t['total'] if t['total'] else 0
                cr = ws2.cell(row=ri, column=3+ci2*2, value=round(rev,2)); sd(cr); cr.number_format='#,##0.00'; cr.fill=fill
                cp = ws2.cell(row=ri, column=4+ci2*2, value=pct); sd(cp); cp.number_format='0.0%'; cp.fill=fill

            # Confidence column
            is_precise = mk >= '2026-03'
            conf_cell = ws2.cell(row=ri, column=len(h2))
            conf_cell.value = '🟢 Server-side' if is_precise else '🟡 Estimat'
            conf_cell.fill = precise_fill if is_precise else est_fill
            sd(conf_cell)

        # Notes section below data
        nr = len(tm) + 4  # leave a gap
        note_font = Font(name='Inter', size=9, italic=True, color='555555')
        note_bold = Font(name='Inter', size=9, bold=True, color='1a1f36')
        notes = [
            "METODOLOGIE & LEGENDĂ — Revenue by Traffic Source",
            "",
            "Total Revenue (coloana B) = total_sales din Shopify, identic cu raportarea internă Zitamine.",
            "Include: toate comenzile (inclusiv TVA, shipping, discount-uri aplicate).",
            "Acest număr coincide exact cu Dashboard-ul intern de vânzări Zitamine.",
            "",
            "LEGENDĂ CANALE:",
            "1. Organic = Căutări naturale sau trafic din postări Social Media nesponsorizate.",
            "   (ex: un utilizator caută pe Google 'vitamine personalizate' și dă click pe rezultatul natural).",
            "2. Paid = Trafic din campanii de publicitate plătite (Meta Ads, Google Ads, TikTok Ads).",
            "   (ex: un utilizator dă click pe o reclamă sponsorizată cu BodyOS pe Facebook).",
            "3. Email (Klaviyo) = Revenue atribuit exclusiv campaniilor de email + flow-urilor automate.",
            "   (ex: un utilizator cumpără direct dintr-un newsletter sau dintr-un email de 'abandoned cart').",
            "4. Direct = Trafic direct, link-uri salvate (bookmarks) sau sursă necunoscută (hidden).",
            "   (ex: reînnoirea automată a abonamentelor prin Appstle, utilizatori cu ad-blockers, link-uri pe WhatsApp).",
            "",
            "Calcul (Mai 2025 - Feb 2026): Restul revenue-ului (Total - Email) a fost distribuit proporțional",
            "între Organic, Paid, Direct și Other conform ponderilor din GA4.",
            "",
            "⚠️ ATENȚIE — Precizia datelor pe canale:",
            "• Mai 2025 – Feb 2026 (🟡 Estimat): GA4 fără server-side tracking captura doar ~32-44%",
            "  din tranzacții. Proporțiile Organic/Paid/Direct sunt APROXIMĂRI bazate pe eșantionul GA4.",
            "  Paid poate fi supraevaluat (tracking mai bun), Direct poate fi subevaluat.",
            "  Abonamentele recurente (Appstle) și comenzile afectate de ad-blockere nu apar în GA4.",
            "",
            "• Mar 2026+ (🟢 Server-side): Server-side tracking activ — GA4 vede ~92-100% din comenzi.",
            "  Datele din această perioadă sunt precise și complete.",
            "",
            "Surse de date: Shopify API (dashboard_data.js), GA4 API (property 297495831), Klaviyo API.",
            f"Generat: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        ]
        for i, note in enumerate(notes):
            cell = ws2.cell(row=nr+i, column=1, value=note)
            ws2.merge_cells(start_row=nr+i, start_column=1, end_row=nr+i, end_column=len(h2))
            cell.font = note_bold if i == 0 or note.startswith('•') or note.startswith('⚠') else note_font
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    for ci, w in enumerate([14,16]+[14,8]*5+[16], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = 'C2'

    # Methodology sheet
    ws_m = wb.create_sheet("Metodologie")
    notes = [
        "METODOLOGIE — Raport Cohorte Abonamente",
        "",
        "Acest raport foloseste o definitie diferita de 'client nou' pentru abonamente.",
        "",
        "OTP (One-Time Purchase):",
        "  • Client nou = prima comandă EVER la Zitamine (nicio comandă anterioară).",
        "  • LTV = Total revenue / # clienți (toate comenzile ulterioare, indiferent de tip).",
        "",
        "SUB1 / SUB3 / SUB6 (Abonamente):",
        "  • Client nou = oricine cu tag appstle_subscription_first_order în luna respectivă.",
        "  • Include și foști clienți OTP care au trecut pe abonament.",
        "  • Retenție = a doua comandă (orice tip) după prima comandă de abonament.",
        "  • LTV = Total revenue de la prima comandă de abonament în sus.",
        "",
        "ALL (Total):",
        "  • Client nou = prima comandă ever (orice tip). Identic cu raportul principal.",
        "",
        "LEGENDĂ COLOANE:",
        "  • Cohortă: Luna în care s-a înregistrat clientul nou (conform definițiilor de mai sus).",
        "  • # Clienți Noi / Noi Abonați: Numărul de clienți din cohortă.",
        "  • % R2 M0-1: Procentul din cohortă care a plasat a 2-a comandă în aceeași lună (M0) sau în luna următoare (M1).",
        "  • % R2 M2: Procentul care a plasat a 2-a comandă în a 2-a lună după prima comandă.",
        "  • % R2 M3, M4-6, etc.: Retenția calculată pe baza lunii în care are loc a 2-a comandă.",
        "  • LTV (Life-Time Value): Total revenue generat de cohortă împărțit la numărul de clienți din cohortă.",
        "",
        "Diferență față de Raportul Principal (Raport_Investitori_Zitamine.xlsx):",
        "  • Raportul principal numără ca SUB doar clienții a căror PRIMĂ comandă a fost abonament.",
        "  • Acest raport numără ca SUB pe oricine care a deschis un abonament nou.",
        "  • Numerele SUB sunt mai mari aici (ex: SUB1 Mai 2025: ~37 vs ~12 în raportul principal).",
        "",
        f"Generat: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        f"Sursă date: master_orders.csv (Shopify API)",
    ]
    note_font = Font(name='Inter', size=10, color='555555')
    note_bold = Font(name='Inter', size=10, bold=True, color='1a1f36')
    for i, note in enumerate(notes):
        cell = ws_m.cell(row=i+1, column=1, value=note)
        ws_m.merge_cells(start_row=i+1, start_column=1, end_row=i+1, end_column=8)
        cell.font = note_bold if i == 0 or note.startswith('OTP') or note.startswith('SUB') or note.startswith('ALL') or note.startswith('Diferență') else note_font
        cell.alignment = Alignment(wrap_text=True, vertical='top')
    ws_m.column_dimensions['A'].width = 90

    os.makedirs(os.path.dirname(OUTPUT_EXCEL), exist_ok=True)
    wb.save(OUTPUT_EXCEL)
    print(f"\nExcel saved: {OUTPUT_EXCEL}")


def main():
    import sys
    sys.path.append(os.path.dirname(__file__))
    try:
        from generate_investor_report import load_traffic
        traffic = load_traffic()
    except Exception as e:
        print(f"Eroare la incarcarea traficului: {e}")
        traffic = None

    orders = load_orders()
    cohorts = build_cohorts(orders)
    gen_excel(cohorts, traffic=traffic)
    print("\nDone!")

if __name__ == '__main__':
    main()
