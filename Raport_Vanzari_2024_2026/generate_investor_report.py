"""Raport Investitori Zitamine - Grupat pe cohorte OTP/SUB1/SUB3/SUB6"""
import csv, json, os
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
MASTER_CSV = os.path.join(SCRIPT_DIR, "master_orders.csv")
DASHBOARD_JS = os.path.join(SCRIPT_DIR, "dashboard_data.js")
GA4_JSON = os.path.join(os.path.dirname(SCRIPT_DIR), "Marketing_Analytics", "ga4_monthly_traffic_extended.json")
OUTPUT_EXCEL = os.path.join(SCRIPT_DIR, "..", "..", "Rapoarte", "Raport_Investitori_Zitamine.xlsx")
OUTPUT_HTML = os.path.join(SCRIPT_DIR, "..", "..", "Rapoarte", "Raport_Investitori_Zitamine.html")
KLAVIYO_JSON = os.path.join(os.path.dirname(SCRIPT_DIR), "Marketing_Analytics", "klaviyo_data.json")

COHORT_START, COHORT_END = "2025-05", "2026-03"
MILESTONES = [1, 2, 3, 6, 9, 12]
MILESTONE_RANGES = {1: (0,1), 2: (2,2), 3: (3,3), 6: (4,6), 9: (7,9), 12: (10,12)}
MILESTONE_LABELS = {1:'M0-1', 2:'M2', 3:'M3', 6:'M4-6', 9:'M7-9', 12:'M10-12'}
REPORT_DATE = datetime(2026, 4, 30)
ORDER_TYPES = ['OTP', 'SUB1', 'SUB3', 'SUB6']
TYPE_COLORS = {'OTP':'#F39C12','SUB1':'#3498DB','SUB3':'#E67E22','SUB6':'#9B59B6'}
TYPE_DESC = {'OTP':'One-Time Purchase','SUB1':'Abonament Lunar','SUB3':'Abonament la 3 Luni','SUB6':'Protocol 6 Luni (rate)'}
ML = {1:'Ian',2:'Feb',3:'Mar',4:'Apr',5:'Mai',6:'Iun',7:'Iul',8:'Aug',9:'Sep',10:'Oct',11:'Nov',12:'Dec'}

def target_months():
    r, y, m = [], 2025, 5
    while f"{y}-{m:02d}" <= COHORT_END:
        r.append(f"{y}-{m:02d}"); m += 1
        if m > 12: m, y = 1, y+1
    return r

def mlabel(mk):
    y, m = mk.split('-'); return f"{ML[int(m)]} {y}"

def parse_date(s):
    if not s: return None
    try:
        c = s.split('+')[0].strip()
        return datetime.strptime(c, '%Y-%m-%dT%H:%M:%S') if 'T' in c else datetime.strptime(c, '%Y-%m-%d')
    except: return None

def classify_type(tags):
    if not tags: return "OTP"
    t = tags.lower()
    if "saseluni" in t: return "SUB6"
    if "treiluni" in t or "treluni" in t: return "SUB3"
    if "appstle_subscription" in t or "subscription" in t: return "SUB1"
    return "OTP"

def mdiff(d1, d2):
    return (d1.year - d2.year)*12 + (d1.month - d2.month)

def load_orders():
    print("Loading master_orders.csv...")
    orders = {}
    with open(MASTER_CSV, 'r', encoding='utf-8-sig') as f:
        for row in csv.DictReader(f):
            name = row.get('Name','').strip()
            if not name or name in orders: continue
            email = row.get('Email','').strip().lower()
            if not email: continue
            if row.get('Cancelled at','').strip(): continue
            fs = (row.get('Financial Status','') or '').strip().lower()
            if fs in ('voided','pending'): continue
            d = parse_date(row.get('Created at',''))
            if not d: continue
            try: total = float(row.get('Total', 0) or 0)
            except: total = 0
            tags = row.get('Tags', '') or ''
            orders[name] = {'email': email, 'date': d, 'total': total,
                           'month': f"{d.year}-{d.month:02d}", 'type': classify_type(tags)}
    print(f"  {len(orders)} valid orders")
    return orders

def build_cohorts(orders):
    tm = target_months()
    cust = defaultdict(list)
    for o in orders.values(): cust[o['email']].append(o)
    for e in cust: cust[e].sort(key=lambda x: x['date'])

    # Structure: cohorts[order_type][month_key] = {customers, total_rev, retained, pct, ltv}
    cohorts = {}
    for ot in ORDER_TYPES:
        cohorts[ot] = {}
        for mk in tm:
            cohorts[ot][mk] = {'label': mlabel(mk), 'customers': 0, 'total_rev': 0,
                               'retained': {ms: 0 for ms in MILESTONES}, 'pct': {ms: None for ms in MILESTONES}}

    # Also build ALL (combined) cohort
    cohorts['ALL'] = {}
    for mk in tm:
        cohorts['ALL'][mk] = {'label': mlabel(mk), 'customers': 0, 'total_rev': 0,
                              'retained': {ms: 0 for ms in MILESTONES}, 'pct': {ms: None for ms in MILESTONES}}

    for email, ords in cust.items():
        first = ords[0]
        cmk = first['month']
        if cmk not in cohorts['ALL']: continue
        first_type = first['type']

        # Add to specific type cohort AND to ALL
        for target in [first_type, 'ALL']:
            cohorts[target][cmk]['customers'] += 1
            cohorts[target][cmk]['total_rev'] += sum(o['total'] for o in ords)
            if len(ords) >= 2:
                m_to_2nd = mdiff(ords[1]['date'], first['date'])
                for ms in MILESTONES:
                    lo, hi = MILESTONE_RANGES[ms]
                    if lo <= m_to_2nd <= hi:
                        cohorts[target][cmk]['retained'][ms] += 1

    for ot in list(cohorts.keys()):
        for mk, c in cohorts[ot].items():
            avail = mdiff(REPORT_DATE, datetime(*map(int, mk.split('-')), 1))
            for ms in MILESTONES:
                if avail >= ms and c['customers'] > 0:
                    c['pct'][ms] = round(c['retained'][ms] / c['customers'] * 100, 1)
            c['ltv'] = round(c['total_rev'] / c['customers'], 2) if c['customers'] > 0 else 0

    for ot in ORDER_TYPES + ['ALL']:
        total = sum(cohorts[ot][mk]['customers'] for mk in tm)
        print(f"  {ot}: {total} customers across {len(tm)} months")
    return cohorts

def load_shopify_revenue():
    """Monthly revenue from dashboard_data.js (total_sales — matches internal reporting)."""
    print("Loading Shopify revenue from dashboard_data.js (internal reporting)...")
    rev = {}
    with open(DASHBOARD_JS, 'r', encoding='utf-8') as f:
        content = f.read()
    import re
    for match in re.finditer(r'"(\d{4}-\d{2})"\s*:\s*\{', content):
        mk = match.group(1)
        ts_match = re.search(r'"total_sales"\s*:\s*([\d.]+)', content[match.start():match.start()+5000])
        if ts_match:
            rev[mk] = float(ts_match.group(1))
    tm = target_months()
    for mk in tm:
        if mk in rev:
            print(f"  {mlabel(mk)}: {rev[mk]:,.0f} RON")
    return rev

def load_klaviyo_revenue():
    """Monthly EMAIL-ONLY revenue from Klaviyo (campaigns by sentDate + flow revenue distributed)."""
    if not os.path.exists(KLAVIYO_JSON): print("  Klaviyo not found"); return {}
    print("Loading Klaviyo email-only revenue...")
    with open(KLAVIYO_JSON, 'r', encoding='utf-8') as f: data = json.load(f)
    
    # 1. Campaign revenue by sent month
    klav = defaultdict(float)
    for c in data.get('campaigns', []):
        rev = c.get('revenue', 0) or 0
        if rev <= 0: continue
        sent = c.get('sentDate', '')
        if sent and len(sent) >= 7:
            mk = sent[:7]
            klav[mk] += rev
    
    # 2. Flow revenue (total across 12 months, distribute evenly)
    flow_total = sum(f.get('revenue', 0) for f in data.get('flows', []))
    tm = target_months()
    flow_per_month = flow_total / max(1, len(tm))
    for mk in tm:
        klav[mk] += flow_per_month
    
    total_email = sum(klav.values())
    print(f"  Campaign revenue: {sum(c.get('revenue',0) for c in data.get('campaigns',[])):,.0f} RON")
    print(f"  Flow revenue: {flow_total:,.0f} RON ({flow_per_month:,.0f}/mo distributed)")
    print(f"  Total email-only: {total_email:,.0f} RON")
    return dict(klav)

def load_traffic():
    """Build hybrid traffic: Shopify total + GA4 % split + Klaviyo email."""
    shopify_rev = load_shopify_revenue()
    klaviyo_rev = load_klaviyo_revenue()
    
    if not os.path.exists(GA4_JSON): print("  GA4 not found"); return None
    print("Loading GA4 traffic split...")
    with open(GA4_JSON, 'r', encoding='utf-8') as f: data = json.load(f)
    
    cat_map = {'Organic Search':'organic','Organic Social':'organic','Organic Shopping':'organic',
               'Paid Search':'paid','Paid Social':'paid','Cross-network':'paid',
               'Paid Other':'paid','Paid Shopping':'paid','Display':'paid',
               'Direct':'direct','Unassigned':'direct','Email':'email'}
    traffic = {}
    for m in data.get('months', []):
        period = m.get('period', ''); start = period.split(' to ')[0] if ' to ' in period else ''
        if not start: continue
        try: d = datetime.strptime(start, '%Y-%m-%d'); mk = f"{d.year}-{d.month:02d}"
        except: continue
        
        # GA4 channel split (for proportions only)
        ga4_split = {'organic':0,'paid':0,'direct':0,'email':0,'other':0}
        ga4_total = 0
        for ch in m.get('channels', []):
            rev = ch.get('totalRevenue', 0) or 0
            cat = cat_map.get(ch.get('sessionDefaultChannelGroup',''), 'other')
            ga4_split[cat] += rev; ga4_total += rev
        
        # Use Shopify as total revenue (source of truth)
        shop_total = shopify_rev.get(mk, 0)
        klav_email = klaviyo_rev.get(mk, 0)
        
        # Email revenue = Klaviyo email-only (campaigns + flows), capped at Shopify total
        email_rev = min(klav_email, shop_total)
        
        # Remaining revenue split by GA4 proportions (exclude email from GA4 split)
        ga4_non_email = ga4_total - ga4_split['email']
        remaining = shop_total - email_rev
        
        row = {'total': shop_total, 'email': round(email_rev, 2)}
        for cat in ['organic', 'paid', 'direct', 'other']:
            if ga4_non_email > 0:
                row[cat] = round(remaining * (ga4_split[cat] / ga4_non_email), 2)
            else:
                row[cat] = round(remaining / 4, 2)
        
        traffic[mk] = row
        print(f"  {mlabel(mk)}: Shopify={shop_total:,.0f} | Email(Klaviyo)={email_rev:,.0f} | Organic={row['organic']:,.0f} | Paid={row['paid']:,.0f}")
    return traffic

def gen_excel(cohorts, traffic):
    tm = target_months()
    wb = Workbook()
    hf = Font(name='Inter', bold=True, color='FFFFFF', size=11)
    hfill = PatternFill(start_color='1a1f36', end_color='1a1f36', fill_type='solid')
    ha = Alignment(horizontal='center', vertical='center', wrap_text=True)
    df = Font(name='Inter', size=10)
    da = Alignment(horizontal='center', vertical='center')
    brd = Border(left=Side('thin','D5D8DC'), right=Side('thin','D5D8DC'),
                 top=Side('thin','D5D8DC'), bottom=Side('thin','D5D8DC'))
    na_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')
    na_font = Font(name='Inter', size=10, color='BDBDBD', italic=True)
    gf = [PatternFill(start_color=c, end_color=c, fill_type='solid')
          for c in ['E8F5E9','C8E6C9','A5D6A7','81C784','66BB6A','4CAF50']]
    type_fills = {t: PatternFill(start_color=c.replace('#',''), end_color=c.replace('#',''), fill_type='solid')
                  for t,c in TYPE_COLORS.items()}

    def sh(ws, row, cols):
        for c in range(1, cols+1):
            cell = ws.cell(row=row, column=c)
            cell.font, cell.fill, cell.alignment, cell.border = hf, hfill, ha, brd

    def sd(cell): cell.font, cell.alignment, cell.border = df, da, brd

    headers = ['Cohortă', '# Clienți Noi\n(prima comandă ever)'] + [f'% R2 {MILESTONE_LABELS[ms]}' for ms in MILESTONES] + ['LTV (RON)']
    ncols = len(headers)

    # Create one sheet per order type + ALL
    for si, ot in enumerate(['ALL'] + ORDER_TYPES):
        if si == 0:
            ws = wb.active; ws.title = "ALL - Total"
        else:
            ws = wb.create_sheet(title=ot)

        # Title row
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
        type_label = f"{ot} — {TYPE_DESC[ot]}" if ot in TYPE_DESC else 'TOTAL (toate tipurile)'
        title_cell = ws.cell(row=1, column=1, value=f"Cohort Retention & LTV — {type_label} — Clienți total noi (prima comandă ever)")
        title_cell.font = Font(name='Inter', bold=True, size=14, color='1a1f36')
        title_cell.alignment = Alignment(horizontal='center', vertical='center')

        # Headers at row 2
        for ci, h in enumerate(headers, 1): ws.cell(row=2, column=ci, value=h)
        sh(ws, 2, ncols)

        for ri, mk in enumerate(tm, 3):
            c = cohorts[ot][mk]
            ws.cell(row=ri, column=1, value=c['label']); sd(ws.cell(row=ri, column=1))
            ws.cell(row=ri, column=1).font = Font(name='Inter', size=10, bold=True)
            ws.cell(row=ri, column=2, value=c['customers']); sd(ws.cell(row=ri, column=2))
            for mi, ms in enumerate(MILESTONES):
                col = 3 + mi; pct = c['pct'][ms]
                cell = ws.cell(row=ri, column=col)
                if pct is not None:
                    cell.value = pct / 100; cell.number_format = '0.0%'; sd(cell)
                    if pct >= 30: cell.fill = gf[min(int(pct/10)-2, 5)]
                    elif pct >= 15: cell.fill = gf[1]
                    elif pct > 0: cell.fill = gf[0]
                else:
                    cell.value = 'N/A'; cell.font, cell.fill = na_font, na_fill
                    cell.alignment, cell.border = da, brd
            ltv_cell = ws.cell(row=ri, column=ncols, value=c['ltv'])
            sd(ltv_cell); ltv_cell.number_format = '#,##0.00 "RON"'

        for ci, w in enumerate([16, 14] + [13]*len(MILESTONES) + [16], 1):
            ws.column_dimensions[get_column_letter(ci)].width = w
        ws.freeze_panes = 'C3'

    # Traffic sheet
    ws2 = wb.create_sheet("Revenue by Traffic Source")
    cats = ['organic','paid','direct','email','other']
    clab = {'organic':'Organic','paid':'Paid','direct':'Direct','email':'Email (Klaviyo)','other':'Other'}
    h2 = ['Luna', 'Total (RON)']
    for cat in cats: h2 += [f'{clab[cat]} (RON)', f'% {clab[cat]}']
    h2.append('Precizie Date')
    for ci, h in enumerate(h2, 1): ws2.cell(row=1, column=ci, value=h)
    sh(ws2, 1, len(h2))
    cfills = {'organic':'E8F5E9','paid':'FFF3E0','direct':'E3F2FD','email':'F3E5F5','other':'F5F5F5'}
    est_fill = PatternFill(start_color='FFF9C4', end_color='FFF9C4', fill_type='solid')  # yellow
    precise_fill = PatternFill(start_color='C8E6C9', end_color='C8E6C9', fill_type='solid')  # green

    if traffic:
        for ri, mk in enumerate(tm, 2):
            t = traffic.get(mk, {k:0 for k in cats+['total']})
            ws2.cell(row=ri, column=1, value=mlabel(mk)); sd(ws2.cell(row=ri, column=1))
            ws2.cell(row=ri, column=1).font = Font(name='Inter', size=10, bold=True)
            ws2.cell(row=ri, column=2, value=round(t['total'],2)); sd(ws2.cell(row=ri, column=2))
            ws2.cell(row=ri, column=2).number_format = '#,##0.00'
            for ci2, cat in enumerate(cats):
                fill = PatternFill(start_color=cfills[cat], end_color=cfills[cat], fill_type='solid')
                rev = t[cat]; pct = rev/t['total'] if t['total'] else 0
                cr = ws2.cell(row=ri, column=3+ci2*2, value=round(rev,2)); sd(cr); cr.number_format='#,##0.00'; cr.fill=fill
                cp = ws2.cell(row=ri, column=4+ci2*2, value=pct); sd(cp); cp.number_format='0.0%'; cp.fill=fill

            # Confidence column
            is_precise = mk >= '2026-03'
            conf_cell = ws2.cell(row=ri, column=len(h2))
            conf_cell.value = '🟢 Server-side' if is_precise else '🟡 Estimat'
            conf_cell.fill = precise_fill if is_precise else est_fill
            sd(conf_cell)

        # Notes section below data
        nr = len(tm) + 4  # leave a gap
        note_font = Font(name='Inter', size=9, italic=True, color='555555')
        note_bold = Font(name='Inter', size=9, bold=True, color='1a1f36')
        notes = [
            "METODOLOGIE & LEGENDĂ — Revenue by Traffic Source",
            "",
            "Total Revenue (coloana B) = total_sales din Shopify, identic cu raportarea internă Zitamine.",
            "Include: toate comenzile (inclusiv TVA, shipping, discount-uri aplicate).",
            "Acest număr coincide exact cu Dashboard-ul intern de vânzări Zitamine.",
            "",
            "LEGENDĂ CANALE:",
            "1. Organic = Căutări naturale sau trafic din postări Social Media nesponsorizate.",
            "   (ex: un utilizator caută pe Google 'vitamine personalizate' și dă click pe rezultatul natural).",
            "2. Paid = Trafic din campanii de publicitate plătite (Meta Ads, Google Ads, TikTok Ads).",
            "   (ex: un utilizator dă click pe o reclamă sponsorizată cu BodyOS pe Facebook).",
            "3. Email (Klaviyo) = Revenue atribuit exclusiv campaniilor de email + flow-urilor automate.",
            "   (ex: un utilizator cumpără direct dintr-un newsletter sau dintr-un email de 'abandoned cart').",
            "4. Direct = Trafic direct, link-uri salvate (bookmarks) sau sursă necunoscută (hidden).",
            "   (ex: reînnoirea automată a abonamentelor prin Appstle, utilizatori cu ad-blockers, link-uri pe WhatsApp).",
            "",
            "Calcul (Mai 2025 - Feb 2026): Restul revenue-ului (Total - Email) a fost distribuit proporțional",
            "între Organic, Paid, Direct și Other conform ponderilor din GA4.",
            "",
            "⚠️ ATENȚIE — Precizia datelor pe canale:",
            "• Mai 2025 – Feb 2026 (🟡 Estimat): GA4 fără server-side tracking captura doar ~32-44%",
            "  din tranzacții. Proporțiile Organic/Paid/Direct sunt APROXIMĂRI bazate pe eșantionul GA4.",
            "  Paid poate fi supraevaluat (tracking mai bun), Direct poate fi subevaluat.",
            "  Abonamentele recurente (Appstle) și comenzile afectate de ad-blockere nu apar în GA4.",
            "",
            "• Mar 2026+ (🟢 Server-side): Server-side tracking activ — GA4 vede ~92-100% din comenzi.",
            "  Datele din această perioadă sunt precise și complete.",
            "",
            "Surse de date: Shopify API (dashboard_data.js), GA4 API (property 297495831), Klaviyo API.",
            f"Generat: {datetime.now().strftime('%d.%m.%Y %H:%M')}",
        ]
        for i, note in enumerate(notes):
            cell = ws2.cell(row=nr+i, column=1, value=note)
            ws2.merge_cells(start_row=nr+i, start_column=1, end_row=nr+i, end_column=len(h2))
            cell.font = note_bold if i == 0 or note.startswith('•') or note.startswith('⚠') else note_font
            cell.alignment = Alignment(wrap_text=True, vertical='top')

    for ci, w in enumerate([14,16]+[14,8]*5+[16], 1):
        ws2.column_dimensions[get_column_letter(ci)].width = w
    ws2.freeze_panes = 'C2'

    ws_m = wb.create_sheet("Metodologie Cohorte")
    notes_coh = [
        "LEGENDĂ TABELE COHORTE",
        "",
        "• Cohortă: Luna în care clientul a plasat prima comandă EVER la Zitamine (orice tip).",
        "• # Clienți Noi: Numărul total de clienți unici din acea cohortă.",
        "• % R2 M0-1: Procentul de clienți care a plasat a 2-a comandă în aceeași lună (M0) sau luna următoare (M1).",
        "• % R2 M2, M3 etc.: Procentul care a plasat a 2-a comandă în a 2-a, a 3-a lună după prima comandă.",
        "• LTV (Life-Time Value): Total revenue generat de clienții din cohortă, împărțit la numărul inițial de clienți.",
        "",
        "Notă: Raportul principal consideră clientul ca fiind 'nou' doar la prima sa tranzacție, indiferent dacă",
        "este OTP sau abonament. (Ex: dacă cineva comandă OTP în Mai, și face abonament în Iulie, el aparține",
        "coortei OTP din Mai, nu apare la clienți noi în Iulie)."
    ]
    for i, note in enumerate(notes_coh):
        cell = ws_m.cell(row=i+1, column=1, value=note)
        ws_m.merge_cells(start_row=i+1, start_column=1, end_row=i+1, end_column=8)
        cell.font = Font(name='Inter', size=10, bold=(i==0), color='1a1f36' if i==0 else '555555')
    ws_m.column_dimensions['A'].width = 100

    os.makedirs(os.path.dirname(OUTPUT_EXCEL), exist_ok=True)
    wb.save(OUTPUT_EXCEL)
    print(f"Excel saved: {OUTPUT_EXCEL}")

def gen_html(cohorts, traffic):
    tm = target_months()
    cats = ['organic','paid','direct','email','other']
    cat_colors = {'organic':'#27ae60','paid':'#e67e22','direct':'#3498db','email':'#9b59b6','other':'#95a5a6'}

    # Build JS data per type
    all_cohort_js = {}
    for ot in ['ALL'] + ORDER_TYPES:
        all_cohort_js[ot] = []
        for mk in tm:
            c = cohorts[ot][mk]
            all_cohort_js[ot].append({'label':c['label'],'customers':c['customers'],'ltv':c['ltv'],
                                      'pct':{str(ms):c['pct'][ms] for ms in MILESTONES}})

    traffic_js = []
    if traffic:
        for mk in tm:
            t = traffic.get(mk, {k:0 for k in cats+['total']})
            traffic_js.append({'label':mlabel(mk), **{k:round(t.get(k,0),2) for k in cats+['total']}})

    s = {'total_customers': sum(cohorts['ALL'][mk]['customers'] for mk in tm),
         'avg_ltv': round(sum(cohorts['ALL'][mk]['ltv'] for mk in tm)/len(tm), 2),
         'total_rev': round(sum(traffic.get(mk,{}).get('total',0) for mk in tm),2) if traffic else 0}
    vals_m3 = [cohorts['ALL'][mk]['pct'][3] for mk in tm if cohorts['ALL'][mk]['pct'][3] is not None]
    s['avg_r2_m3'] = round(sum(vals_m3)/max(1,len(vals_m3)), 1) if vals_m3 else 0

    def cohort_table(ot, data):
        type_label = f"{ot} — {TYPE_DESC[ot]}" if ot in TYPE_DESC else '🔵 TOTAL (toate tipurile)'
        h = f"<h3 style='margin:16px 0 8px;color:{TYPE_COLORS.get(ot,'#1a1f36')}'>{type_label}</h3>"
        h += "<p style='font-size:.78rem;color:#697386;margin-bottom:8px'>⚠️ Clienți total noi — prima comandă ever la Zitamine</p>"
        h += "<table><thead><tr><th>Cohortă</th><th># Noi</th>"
        for ms in MILESTONES: h += f"<th>{MILESTONE_LABELS[ms]}</th>"
        h += "<th>LTV</th></tr></thead><tbody>"
        for c in data:
            h += f"<tr><td style='font-weight:600;text-align:left'>{c['label']}</td><td>{c['customers']}</td>"
            for ms in MILESTONES:
                p = c['pct'][str(ms)]
                if p is None: h += "<td class='na'>N/A</td>"
                else:
                    cls = 'pct-high' if p>=25 else 'pct-mid' if p>=10 else 'pct-low' if p>0 else ''
                    h += f"<td class='{cls}'>{p:.1f}%</td>"
            h += f"<td class='ltv'>{c['ltv']:,.0f}</td></tr>"
        h += "</tbody></table>"
        return h

    sections_html = ""
    for ot in ['ALL'] + ORDER_TYPES:
        sections_html += cohort_table(ot, all_cohort_js[ot])

    traffic_html = ""
    if traffic_js:
        traffic_html = "<table><thead><tr><th>Luna</th><th>Total</th><th>Organic</th><th>Paid</th><th>Direct</th><th>Email</th><th>Other</th></tr></thead><tbody>"
        for t in traffic_js:
            traffic_html += f"<tr><td style='font-weight:600;text-align:left'>{t['label']}</td><td>{t['total']:,.0f}</td>"
            for cat in cats:
                pct = t[cat]/t['total']*100 if t['total'] else 0
                traffic_html += f"<td>{t[cat]:,.0f} <small style='color:#999'>({pct:.0f}%)</small></td>"
            traffic_html += "</tr>"
        traffic_html += "</tbody></table>"

    html = f"""<!DOCTYPE html><html lang="ro"><head>
<meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1">
<title>Zitamine — Investor Report</title>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@300;400;500;600;700&display=swap" rel="stylesheet">
<script src="https://cdn.jsdelivr.net/npm/chart.js@4"></script>
<style>
*{{margin:0;padding:0;box-sizing:border-box}}
body{{font-family:'Inter',sans-serif;background:#f8f9fc;color:#1a1f36;min-height:100vh}}
.container{{max-width:1400px;margin:0 auto;padding:32px 24px}}
header{{text-align:center;margin-bottom:40px}}
header h1{{font-size:2.2rem;font-weight:700;background:linear-gradient(135deg,#1a1f36,#635bff);-webkit-background-clip:text;-webkit-text-fill-color:transparent}}
header .sub{{color:#697386;font-size:1rem;margin-top:4px}}
header .period{{display:inline-block;margin-top:12px;padding:6px 20px;background:#635bff;color:#fff;border-radius:20px;font-size:.85rem;font-weight:500}}
.kpis{{display:grid;grid-template-columns:repeat(4,1fr);gap:20px;margin-bottom:40px}}
.kpi{{background:#fff;border-radius:16px;padding:24px;box-shadow:0 1px 3px rgba(0,0,0,.06);text-align:center}}
.kpi .val{{font-size:1.8rem;font-weight:700;color:#1a1f36}}
.kpi .lbl{{font-size:.8rem;color:#697386;text-transform:uppercase;letter-spacing:.5px;margin-top:4px}}
section{{background:#fff;border-radius:16px;padding:28px;box-shadow:0 1px 3px rgba(0,0,0,.06);margin-bottom:28px}}
section h2{{font-size:1.3rem;font-weight:600;margin-bottom:20px;color:#1a1f36}}
table{{width:100%;border-collapse:collapse;font-size:.82rem;margin-bottom:12px}}
th{{background:#1a1f36;color:#fff;padding:8px 6px;text-align:center;font-weight:500;font-size:.78rem}}
td{{padding:6px;text-align:center;border-bottom:1px solid #e8e8ee}}
tr:hover td{{background:#f5f6fa}}
.na{{color:#ccc;font-style:italic}}.ltv{{font-weight:600;color:#635bff}}
.pct-high{{background:#e8f5e9;color:#2e7d32;font-weight:600}}
.pct-mid{{background:#fff8e1;color:#f57f17}}.pct-low{{background:#fce4ec;color:#c62828}}
.charts{{display:grid;grid-template-columns:1fr 1fr;gap:24px;margin-top:24px}}
.chart-box{{background:#fafbfc;border-radius:12px;padding:16px}}
canvas{{max-height:300px}}
.legend{{background:#f5f6fa;padding:16px;border-radius:8px;font-size:.82rem;color:#444;margin-bottom:24px;border-left:4px solid #635bff}}
.legend h4{{margin:0 0 8px;font-size:.9rem;color:#1a1f36}}
.legend ul{{margin:0;padding-left:20px;line-height:1.5}}
footer{{text-align:center;color:#aaa;font-size:.75rem;margin-top:40px;padding:20px}}
@media(max-width:900px){{.kpis,.charts{{grid-template-columns:1fr 1fr}}}}
</style></head><body><div class="container">
<header><h1>Zitamine — Investor Report</h1>
<p class="sub">Cohort Retention & LTV — Grupat pe OTP / SUB1 / SUB3 / SUB6</p>
<span class="period">Mai 2025 — Aprilie 2026</span></header>
<div class="kpis">
<div class="kpi"><div class="val">{s['total_customers']:,}</div><div class="lbl">Total Clienți Noi</div></div>
<div class="kpi"><div class="val">{s['avg_ltv']:.0f} RON</div><div class="lbl">LTV Mediu</div></div>
<div class="kpi"><div class="val">{s['avg_r2_m3']:.1f}%</div><div class="lbl">Avg Retention ≤M3</div></div>
<div class="kpi"><div class="val">{s['total_rev']:,.0f} RON</div><div class="lbl">Revenue Total (GA4)</div></div>
</div>
<section><h2>📊 Cohort Retention & LTV per Tip Comandă</h2>
<div class="legend">
    <h4>Legendă Tabel Cohorte</h4>
    <ul>
        <li><b>Cohortă:</b> Luna în care clientul a plasat prima comandă <i>ever</i> la Zitamine.</li>
        <li><b># Noi:</b> Numărul total de clienți unici din cohortă.</li>
        <li><b>% R2 M0-1:</b> Procentul din cohortă care a plasat a 2-a comandă în aceeași lună (M0) sau în luna imediat următoare (M1).</li>
        <li><b>% R2 M2, M3 etc.:</b> Procentul din cohortă care a plasat a 2-a comandă în a 2-a, a 3-a lună, etc. după prima comandă.</li>
        <li><b>LTV:</b> Life-Time Value mediu. Total revenue generat de clienții din cohortă, împărțit la numărul inițial de clienți.</li>
    </ul>
</div>
{sections_html}
<div class="charts"><div class="chart-box"><canvas id="ltvChart"></canvas></div>
<div class="chart-box"><canvas id="retChart"></canvas></div></div></section>
<section><h2>📈 Revenue by Traffic Source</h2>
<div class="legend" style="border-left-color:#27ae60">
    <h4>Legendă Canale de Trafic</h4>
    <ul>
        <li><b>Total:</b> total_sales din Shopify (inclusiv TVA, shipping), identic cu Dashboard-ul intern.</li>
        <li><b>Organic:</b> Căutări naturale sau trafic din postări Social Media nesponsorizate.</li>
        <li><b>Paid:</b> Trafic din campanii de publicitate plătite (Meta Ads, Google Ads).</li>
        <li><b>Direct:</b> Utilizatori care scriu adresa direct, folosesc bookmarks sau reînnoiri automate de abonamente (Appstle).</li>
        <li><b>Email:</b> Revenue atribuit exclusiv campaniilor de email și flow-urilor automate (din Klaviyo API).</li>
    </ul>
</div>
{traffic_html}
<div class="charts"><div class="chart-box"><canvas id="trafficChart"></canvas></div>
<div class="chart-box"><canvas id="trafficPctChart"></canvas></div></div></section>
<footer>Generat automat — {datetime.now().strftime('%d.%m.%Y %H:%M')} | Zitamine SRL</footer>
</div><script>
const D={json.dumps(all_cohort_js)};
const T={json.dumps(traffic_js)};
const TC={json.dumps(TYPE_COLORS)};
const types=['OTP','SUB1','SUB3','SUB6'];
// LTV comparison chart
new Chart(document.getElementById('ltvChart'),{{type:'bar',data:{{
labels:D.ALL.map(c=>c.label),
datasets:types.map(t=>({{label:t,data:D[t].map(c=>c.ltv),backgroundColor:TC[t],borderRadius:4}}))
}},options:{{responsive:true,plugins:{{title:{{display:true,text:'LTV per Cohortă per Tip'}}}}}}}});
// Retention M3 comparison
new Chart(document.getElementById('retChart'),{{type:'bar',data:{{
labels:D.ALL.map(c=>c.label),
datasets:types.map(t=>({{label:t+' ≤M3',data:D[t].map(c=>c.pct['3']),backgroundColor:TC[t],borderRadius:4}}))
}},options:{{responsive:true,plugins:{{title:{{display:true,text:'Retention ≤M3 per Tip'}}}}}}}});
if(T.length){{
const cats=['organic','paid','direct','email','other'];
const cc={json.dumps(cat_colors)};
const cl={{organic:'Organic',paid:'Paid',direct:'Direct',email:'Email',other:'Other'}};
new Chart(document.getElementById('trafficChart'),{{type:'bar',data:{{
labels:T.map(t=>t.label),datasets:cats.map(c=>({{label:cl[c],data:T.map(t=>t[c]),backgroundColor:cc[c]}}))
}},options:{{responsive:true,scales:{{x:{{stacked:true}},y:{{stacked:true}}}},plugins:{{title:{{display:true,text:'Revenue by Source (RON)'}}}}}}}});
new Chart(document.getElementById('trafficPctChart'),{{type:'bar',data:{{
labels:T.map(t=>t.label),datasets:cats.map(c=>({{label:cl[c],data:T.map(t=>t.total?t[c]/t.total*100:0),backgroundColor:cc[c]}}))
}},options:{{responsive:true,scales:{{x:{{stacked:true}},y:{{stacked:true,max:100}}}},plugins:{{title:{{display:true,text:'Revenue Split (%)'}}}}}}}});
}}
</script></body></html>"""
    os.makedirs(os.path.dirname(OUTPUT_HTML), exist_ok=True)
    with open(OUTPUT_HTML, 'w', encoding='utf-8') as f: f.write(html)
    print(f"HTML saved: {OUTPUT_HTML}")

if __name__ == '__main__':
    orders = load_orders()
    cohorts = build_cohorts(orders)
    traffic = load_traffic()
    gen_excel(cohorts, traffic)
    gen_html(cohorts, traffic)
    print("\nDone!")
